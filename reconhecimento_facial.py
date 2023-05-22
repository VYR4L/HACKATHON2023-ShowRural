import os
import cv2
import threading
import tkinter as tk
import face_recognition
from datetime import datetime
from PIL import Image, ImageTk
from openpyxl import load_workbook


def exibir_frame():
    ret, frame = webcam.read()
    if ret:
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        imagem = Image.fromarray(frame_rgb)
        imagem = ImageTk.PhotoImage(imagem)
        label_imagem.configure(image=imagem)
        label_imagem.image = imagem
    janela.after(10, exibir_frame)
    janela.update()


fim_do_prrograma = False
janela = tk.Tk()
janela.title("Interface Gráfica")
largura_tela = janela.winfo_screenwidth()
altura_tela = janela.winfo_screenheight()
janela.geometry(f"{largura_tela}x{altura_tela}")

frame_webcam = tk.Frame(janela, width=480, height=360)
label_imagem = tk.Label(frame_webcam)
label_imagem.pack()

frame_webcam.place(x=0, y=0)


def fechar_programa():
    global encerrar_programa
    encerrar_programa = True
    janela.quit()
    webcam.release()
    cv2.destroyAllWindows()


botao_fechar = tk.Button(janela, text="Fechar", command=fechar_programa)
botao_fechar.pack()
botao_fechar.place(x=240, y=500)


webcam = cv2.VideoCapture(0)
encerrar_programa = False


def camera():
    nao_reconhece = 0
    global encerrar_programa
    wbf = load_workbook('/home/vyral/Vídeos/Face_recognizer/Funcionarios.xlsx')
    wbc = load_workbook('/home/vyral/Vídeos/Face_recognizer/Convidados.xlsx')

    wsf = wbf.active
    wsc = wbc.active

    linha_funcionarios_xlsx = wsf.max_row + 1
    linha_convidados_xlsx = wsc.max_row + 1

    imagens_referencia = []

    lista_referenca_funcionarios = []
    lista_referenca_convidados = []

    while not encerrar_programa:

        # Adiciona as imagens de funcionários à lista de imagens de referência
        for arquivo in os.listdir("/home/vyral/Vídeos/Face_recognizer/funcionarios"):
            if arquivo.endswith(".jpg"):
                nome_sem_extensao = os.path.splitext(arquivo)[0]
                imagem_referencia = face_recognition.load_image_file(os.path.join("/home/vyral/Vídeos/Face_recognizer/funcionarios", arquivo))
                features_referencia = face_recognition.face_encodings(imagem_referencia)[0]
                lista_referenca_funcionarios.append(nome_sem_extensao)
                imagens_referencia.append(features_referencia)

        # Adiciona as imagens de convidados à lista de imagens de referência
        for arquivo in os.listdir("/home/vyral/Vídeos/Face_recognizer/convidados"):
            if arquivo.endswith(".jpg"):
                nome_sem_extensao = os.path.splitext(arquivo)[0]
                imagem_referencia = face_recognition.load_image_file(os.path.join("/home/vyral/Vídeos/Face_recognizer/convidados", arquivo))
                features_referencia = face_recognition.face_encodings(imagem_referencia)[0]
                lista_referenca_convidados.append(nome_sem_extensao)
                imagens_referencia.append(features_referencia)
                nao_reconhece += 1

        lista_referenca_funcionarios.reverse()
        lista_referenca_convidados.reverse()

        validacao, frame = webcam.read()

        imagem = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        rostos = face_recognition.face_locations(imagem)
        features_rostos = face_recognition.face_encodings(imagem, rostos)

        rosto_reconhecido = False

        for features_rosto in features_rostos:
            for features_referencia in imagens_referencia:
                distancia = face_recognition.face_distance([features_referencia], features_rosto)
                if distancia < 0.5:
                    rosto_reconhecido = True
                    nome_funcionario_reconhecido = nome_sem_extensao
                    if rosto_reconhecido and nome_sem_extensao in lista_referenca_funcionarios:
                        wsf['A' + str(linha_funcionarios_xlsx)] = nome_funcionario_reconhecido
                        wsf['B' + str(linha_funcionarios_xlsx)] = datetime.now()
                        linha_funcionarios_xlsx += 1
                        wbf.save('/home/vyral/Vídeos/Face_recognizer/Funcionarios.xlsx')
                        lista_referenca_funcionarios.remove(nome_funcionario_reconhecido)

            if not rosto_reconhecido:
                nao_reconhece += 1
                global teste
                teste = True
                for index in range(nao_reconhece):
                    ret, frame_convidado = webcam.read()
                    cv2.imwrite(f"/home/vyral/Vídeos/Face_recognizer/convidados/convidado{index}.jpg", frame_convidado)
                    wsc['A' + str(linha_convidados_xlsx)] = f'convidado{index}'
                    wsc['B' + str(linha_convidados_xlsx)] = datetime.now()
                    linha_convidados_xlsx += 1
                    wbc.save('/home/vyral/Vídeos/Face_recognizer/Convidados.xlsx')
                    lista_referenca_convidados.remove(f'convidado{index}')


face_thread = threading.Thread(target=camera)
