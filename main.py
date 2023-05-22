from reconhecimento_facial import face_thread, exibir_frame, janela
from openpyxl import load_workbook


wbf = load_workbook('/home/vyral/Vídeos/Face_recognizer/Funcionarios.xlsx')
wbc = load_workbook('/home/vyral/Vídeos/Face_recognizer/Convidados.xlsx')

wsf = wbf.active
wsc = wbc.active

wsf.delete_rows(1, wsf.max_row)
wsc.delete_rows(1, wsc.max_row)

wbf.save('/home/vyral/Vídeos/Face_recognizer/Funcionarios.xlsx')
wbc.save('/home/vyral/Vídeos/Face_recognizer/Convidados.xlsx')

face_thread.start()
exibir_frame()
janela.mainloop()

wbf.close()
wbc.close()
